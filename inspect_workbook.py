import json
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter


path = Path(r"C:\Users\admin\Desktop\excel tool\折旧表_脱敏.xlsx")
wb_formula = load_workbook(path, data_only=False)
wb_value = load_workbook(path, data_only=True)

result = {
    "workbook": {
        "sheets": wb_formula.sheetnames,
        "defined_names": [str(x) for x in wb_formula.defined_names.values()],
        "calculation": str(wb_formula.calculation),
    },
    "sheets": [],
}

for ws in wb_formula.worksheets:
    wsv = wb_value[ws.title]
    cells = []
    formulas = []
    styles = {}
    for row in ws.iter_rows():
        row_values = []
        has_value = False
        for cell in row:
            value = cell.value
            if value is not None:
                has_value = True
            row_values.append(value)
            if isinstance(value, str) and value.startswith("="):
                formulas.append(
                    {
                        "cell": cell.coordinate,
                        "formula": value,
                        "cached": wsv[cell.coordinate].value,
                    }
                )
            if value is not None:
                style_key = (
                    cell.style_id,
                    cell.number_format,
                    cell.alignment.horizontal,
                    cell.alignment.vertical,
                    cell.fill.fgColor.rgb,
                    cell.font.name,
                    cell.font.sz,
                    cell.font.bold,
                )
                styles.setdefault(str(style_key), []).append(cell.coordinate)
        if has_value:
            cells.append({"row": row[0].row, "values": row_values})

    validations = []
    if ws.data_validations:
        for dv in ws.data_validations.dataValidation:
            validations.append(
                {
                    "type": dv.type,
                    "sqref": str(dv.sqref),
                    "formula1": dv.formula1,
                    "formula2": dv.formula2,
                }
            )

    result["sheets"].append(
        {
            "title": ws.title,
            "dimensions": ws.calculate_dimension(),
            "max_row": ws.max_row,
            "max_column": ws.max_column,
            "sheet_state": ws.sheet_state,
            "freeze_panes": str(ws.freeze_panes) if ws.freeze_panes else None,
            "merged_cells": [str(x) for x in ws.merged_cells.ranges],
            "auto_filter": str(ws.auto_filter.ref) if ws.auto_filter.ref else None,
            "tables": list(ws.tables.keys()),
            "data_validations": validations,
            "hidden_rows": [
                index for index, dim in ws.row_dimensions.items() if dim.hidden
            ],
            "hidden_columns": [
                key for key, dim in ws.column_dimensions.items() if dim.hidden
            ],
            "row_heights": {
                str(index): dim.height
                for index, dim in ws.row_dimensions.items()
                if dim.height is not None
            },
            "column_widths": {
                key: dim.width
                for key, dim in ws.column_dimensions.items()
                if dim.width is not None
            },
            "print": {
                "orientation": ws.page_setup.orientation,
                "paper_size": ws.page_setup.paperSize,
                "fit_to_width": ws.page_setup.fitToWidth,
                "fit_to_height": ws.page_setup.fitToHeight,
                "print_area": str(ws.print_area),
                "print_title_rows": str(ws.print_title_rows),
                "margins": {
                    "left": ws.page_margins.left,
                    "right": ws.page_margins.right,
                    "top": ws.page_margins.top,
                    "bottom": ws.page_margins.bottom,
                    "header": ws.page_margins.header,
                    "footer": ws.page_margins.footer,
                },
            },
            "sheet_view": {
                "show_grid_lines": ws.sheet_view.showGridLines,
                "zoom_scale": ws.sheet_view.zoomScale,
                "zoom_scale_normal": ws.sheet_view.zoomScaleNormal,
            },
            "cells": cells,
            "formulas": formulas,
            "style_groups": styles,
        }
    )

print(json.dumps(result, ensure_ascii=False, indent=2, default=str))
