import hashlib
import json
import zipfile
from pathlib import Path

from openpyxl import load_workbook


root = Path(r"C:\Users\admin\Desktop\excel tool")
tool = root / "月度折旧生成工具.xlsm"
source = root / "折旧表_脱敏.xlsx"
wps_output = Path(r"C:\tmp\depreciation_qa_v12\new_complete_auto_dep\折旧表_2026.02.xlsx")
excel_output = Path(r"C:\tmp\depreciation_qa_excel16\new_complete_auto_dep\折旧表_2026.02.xlsx")


def sha256(path):
    h = hashlib.sha256()
    with path.open("rb") as f:
        for chunk in iter(lambda: f.read(1024 * 1024), b""):
            h.update(chunk)
    return h.hexdigest().upper()


with zipfile.ZipFile(tool) as zf:
    names = set(zf.namelist())
    xml_text = "\n".join(
        zf.read(name).decode("utf-8", errors="ignore")
        for name in names
        if name.endswith((".xml", ".rels", ".vml"))
    )
    macro_hits = {
        macro: macro in xml_text
        for macro in (
            "SelectSource",
            "CheckAndPreview",
            "GenerateMonth",
            "ClearAll",
        )
    }
    has_vba = "xl/vbaProject.bin" in names

wb = load_workbook(tool, keep_vba=True, data_only=False)
home = wb["操作首页"]
headers = [home.cell(10, c).value for c in range(1, 14)]
rules = [home["J5"].value, home["J6"].value, home["J7"].value]
validations = [
    {"sqref": str(dv.sqref), "type": dv.type, "formula1": dv.formula1}
    for dv in home.data_validations.dataValidation
]

def validate_output(path):
    with zipfile.ZipFile(path) as zf:
        has_output_vba = "xl/vbaProject.bin" in zf.namelist()
    out_wb = load_workbook(path, data_only=False)
    month = out_wb["2026.02"]
    monthly_formula_errors = []
    for row in month.iter_rows(
        min_row=1, max_row=month.max_row, min_col=1, max_col=16
    ):
        for cell in row:
            if isinstance(cell.value, str) and any(
                err in cell.value
                for err in ("#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A")
            ):
                monthly_formula_errors.append(f"{cell.coordinate}:{cell.value}")
    return {
        "has_vba_project": has_output_vba,
        "sheets": out_wb.sheetnames,
        "freeze": str(month.freeze_panes),
        "orientation": month.page_setup.orientation,
        "formula_errors": monthly_formula_errors,
        "new_asset": {
            "asset_no": month["C4"].value,
            "category": month["A4"].value,
            "unit": month["D4"].value,
            "quantity": month["E4"].value,
            "monthly": month["M4"].value,
            "depreciation": month["N4"].value,
            "ending_net": month["O4"].value,
        },
        "total_formulas": [month["N5"].value, month["O5"].value],
    }

print(
    json.dumps(
        {
            "tool_sha256": sha256(tool),
            "source_sha256": sha256(source),
            "tool_has_vba_project": has_vba,
            "button_macro_bindings_found": macro_hits,
            "home_headers": headers,
            "home_rules": rules,
            "home_print_area": str(home.print_area),
            "home_validations": validations,
            "wps_output": validate_output(wps_output),
            "excel16_output": validate_output(excel_output),
        },
        ensure_ascii=True,
        indent=2,
    )
)
