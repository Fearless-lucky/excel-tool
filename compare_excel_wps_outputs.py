import json
from pathlib import Path

from openpyxl import load_workbook


wps_path = Path(r"C:\tmp\depreciation_qa_v12\new_complete_auto_dep\折旧表_2026.02.xlsx")
excel_path = Path(r"C:\tmp\depreciation_qa_excel16\new_complete_auto_dep\折旧表_2026.02.xlsx")


def snapshot(path):
    wb = load_workbook(path, data_only=False)
    ws = wb["2026.02"]
    cell_data = {}

    def color_value(color):
        if color is None:
            return None
        return (
            color.type,
            str(color.rgb) if color.rgb is not None else None,
            str(color.indexed) if color.indexed is not None else None,
            str(color.theme) if color.theme is not None else None,
            str(color.tint) if color.tint is not None else None,
        )

    def border_value(border):
        return tuple(
            (getattr(border, side).style, color_value(getattr(border, side).color))
            for side in ("left", "right", "top", "bottom")
        )

    def number_format_class(fmt):
        if "#,##0.00" in fmt and '"-"' in fmt:
            return "accounting_2"
        if fmt in ("0.00_ ", "0.00"):
            return "decimal_2"
        return fmt

    for row in ws.iter_rows(min_row=1, max_row=ws.max_row, min_col=1, max_col=16):
        for cell in row:
            if cell.value is not None:
                cell_data[cell.coordinate] = {
                    "value": str(cell.value),
                    "number_format": number_format_class(cell.number_format),
                    "font": (
                        cell.font.name,
                        cell.font.sz,
                        cell.font.bold,
                        cell.font.italic,
                        color_value(cell.font.color),
                    ),
                    "fill": (
                        cell.fill.fill_type,
                        color_value(cell.fill.fgColor),
                        color_value(cell.fill.bgColor),
                    ),
                    "border": border_value(cell.border),
                    "alignment": (
                        cell.alignment.horizontal,
                        cell.alignment.vertical,
                        cell.alignment.wrap_text,
                        cell.alignment.shrink_to_fit,
                        cell.alignment.text_rotation,
                    ),
                }
    return {
        "sheets": wb.sheetnames,
        "dimension": ws.calculate_dimension(),
        "cells": cell_data,
        "merged": sorted(str(x) for x in ws.merged_cells.ranges),
        "freeze": str(ws.freeze_panes),
        "orientation": ws.page_setup.orientation,
        "paper_size": ws.page_setup.paperSize,
        "print_area": str(ws.print_area),
        "print_title_rows": str(ws.print_title_rows),
        "column_widths": {
            key: dim.width
            for key, dim in ws.column_dimensions.items()
            if dim.width is not None and len(key) <= 2 and key <= "P"
        },
        "row_heights": {
            str(key): dim.height
            for key, dim in ws.row_dimensions.items()
            if dim.height is not None
        },
        "hidden_columns": sorted(
            key for key, dim in ws.column_dimensions.items() if dim.hidden
        ),
    }


wps = snapshot(wps_path)
excel = snapshot(excel_path)
differences = {}
for key in wps:
    if key == "column_widths":
        width_keys = set(wps[key]) | set(excel[key])
        width_diff = {
            column: {"wps": wps[key].get(column), "excel": excel[key].get(column)}
            for column in width_keys
            if abs((wps[key].get(column) or 0) - (excel[key].get(column) or 0)) > 0.05
        }
        if width_diff:
            differences[key] = width_diff
    elif wps[key] != excel[key]:
        differences[key] = {"wps": wps[key], "excel": excel[key]}

print(
    json.dumps(
        {
            "equivalent": not differences,
            "difference_keys": list(differences),
            "differences": differences,
        },
        ensure_ascii=True,
        indent=2,
    )
)
